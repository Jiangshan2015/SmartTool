from typing import Union
import flet as ft
from route.Router import Router
from time import sleep
import openpyxl
import os
from openpyxl_image_loader import SheetImageLoader
from tool.smart_tool import SmartTool as sl


def ReadFileView(router: Union[Router, str, None] = None):
    pb = ft.ProgressBar(width=400, visible=False)
    dir_pre = "文件路径："
    directory_path = ft.Text(dir_pre)
    progress_state = ft.Text("")

    # 返回
    def goback(e: ft.ControlEvent):
        close_banner(e)
        router.page.controls.pop()
        e.page.go("/")

    # Open directory dialog
    def get_directory_result(e: ft.FilePickerResultEvent):
        directory_path.value = dir_pre + e.path if e.path else ""
        directory_path.update()

    def close_banner(e):
        router.page.banner.open = False
        router.page.update()

    router.page.banner = ft.Banner(
        bgcolor=ft.colors.AMBER_100,
        leading=ft.Icon(ft.icons.WARNING_AMBER_ROUNDED, color=ft.colors.AMBER, size=20),
        content=ft.Text(
            "ooo，你没有选择正确文件夹，重试！"
        ),
        actions=[
            ft.TextButton(" "),
            ft.TextButton("忽略", on_click=close_banner),
        ],
    )

    def handleFiles():
        save_dir = sl.makeDir('Smart_tool_images')
        print(save_dir)
        dir_path = directory_path.value.removeprefix(dir_pre)
        file_paths = []
        for root, dirs, files in os.walk(dir_path):
            file_paths = [os.path.join(root, file) for file in files]
        for file_path in file_paths:
            if file_path.endswith('.xlsx'):
                print(file_path)
                workbook = openpyxl.load_workbook(file_path)
                for sheetName in workbook.sheetnames:
                    ws = workbook[sheetName]
                    image_loader = SheetImageLoader(ws)
                    num = ws.max_row

                    for i in range(2, num + 1):  # 从第2行开始，总行数要+1
                        try:
                            name = ws['B' + str(i)].value  # B列的文件名
                            image = image_loader.get('C' + str(i))  # C列的图片
                            image = image.convert('RGB')
                            image_name = save_dir + '/' + name + ".jpg"
                            image.save(image_name)  # 以Ai为名，存图片Ci
                        # 排除没有图片，或图片超出单元格的情况
                        except ValueError:
                            print("这一行没有图片：", i)

    def read_file(e: ft.FilePickerResultEvent):
        router.page.banner.open = False
        router.page.update()
        if directory_path.value == dir_pre:
            router.page.banner.open = True
            router.page.update()
            return
        pb.visible = True
        handleFiles()
        progress_state.value = "正在读取文件，稍后..."
        router.page.update()
        sleep(1)
        progress_state.value = "读取完成"
        pb.value = 1
        router.page.update()

    get_directory_dialog = ft.FilePicker(on_result=get_directory_result)

    router.page.overlay.extend([get_directory_dialog])
    router.page.add(
        ft.Container(
            margin=ft.margin.only(top=20),
            content=ft.Row(
                [
                    ft.FilledButton("返回首页", icon="ARROW_BACK_IOS_OUTLINED", on_click=goback),
                    ft.Card(

                        width=600,
                        height=450,
                        content=ft.Container(
                            image_src=f"../assets/bg_content.jpeg",
                            image_fit=ft.ImageFit.COVER,
                            border_radius=5,
                            content=ft.Column(
                                [
                                    ft.Container(
                                        ft.Text("备注.生成的图片在桌面Smart_tool_images文件夹中", size=14),
                                        width=600,
                                        height=40,
                                        padding=10
                                    ),
                                    ft.Container(
                                        width=600,
                                        height=50,
                                        alignment=ft.alignment.center,
                                        content=ft.FilledButton(
                                            "打开文件夹",
                                            icon=ft.icons.FOLDER_OPEN,
                                            on_click=lambda _: get_directory_dialog.get_directory_path(),
                                        ),
                                    ),
                                    # ft.icons.UPLOAD
                                    ft.Container(
                                        width=600,
                                        height=70,
                                        padding=10,
                                        alignment=ft.alignment.center,
                                        content=directory_path,
                                    ),
                                    ft.FilledButton("开始读取图片", icon="UPLOAD", on_click=read_file),
                                    ft.Container(
                                        width=600,
                                        height=50,
                                        alignment=ft.alignment.center,
                                        content=ft.Column([progress_state, pb]),
                                    ),

                                ],
                                wrap=True,
                                spacing=30,
                                run_spacing=10,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            )
                        )),
                ],
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.START
            )
        ),
    )
