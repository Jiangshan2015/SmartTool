from typing import Union
import flet as ft
from route.Router import Router
from time import sleep
import os
import openpyxl
from openpyxl.drawing.image import Image


def WriteFileView(router: Union[Router, str, None] = None):
    pb = ft.ProgressBar(width=400, visible=False)
    dir_pre = '写入文件路径：'
    directory_path = ft.Text(dir_pre)
    progress_state = ft.Text("")

    # 返回
    def goback(e: ft.ControlEvent):
        close_banner(e)
        router.page.controls.pop()
        e.page.go("/")

    # Open directory dialog
    def get_directory_result(e: ft.FilePickerResultEvent):
        directory_path.value = dir_pre + e.path if e.path else ""
        directory_path.update()

    def close_banner(e):
        router.page.banner.open = False
        router.page.update()

    router.page.banner = ft.Banner(
        bgcolor=ft.colors.AMBER_100,
        leading=ft.Icon(ft.icons.WARNING_AMBER_ROUNDED, color=ft.colors.AMBER, size=20),
        content=ft.Text(
            "ooo，你没有选择正确文件夹，重试！"
        ),
        actions=[
            ft.TextButton(" "),
            ft.TextButton("忽略", on_click=close_banner),
        ],
    )

    def handleFiles():
        dir_path = directory_path.value.removeprefix(dir_pre)
        file_paths = []
        img_name_column = 'B'
        img_column = 'C'
        img_dir = os.path.expanduser('~/Desktop/Smart_tool_images')
        for root, dirs, files in os.walk(dir_path):
            file_paths = [os.path.join(root, file) for file in files]
        for file_path in file_paths:
            if file_path.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file_path)
                # 获取sheet页
                for sheetName in workbook.sheetnames:
                    ws = workbook[sheetName]
                    num = ws.max_row
                    for i in range(2, num + 1):  # 从第2行开始，总行数要+1
                        img_file_path = ''
                        img_name = ws[img_name_column + str(i)].value
                        if img_name is not None:
                            img_file_path = img_dir + '/' + img_name + '.jpg'
                        if os.path.exists(img_file_path):
                            # 获取图片
                            img = Image(img_file_path)
                            # 设置图片的大小
                            img.width, img.height = (110, 110)
                            # # 设置表格的宽20和高85
                            # ws.column_dimensions[img_column].width = 20
                            # ws.row_dimensions[i].height = 85
                            # 图片插入名称对应单元格
                            ws.add_image(img, anchor=img_column + str(i))
                workbook.save(file_path)
                workbook.close()
                print(f'保存完成')

    def write_file(e: ft.FilePickerResultEvent):
        router.page.banner.open = False
        router.page.update()
        if directory_path.value == dir_pre:
            router.page.banner.open = True
            router.page.update()
            return
        handleFiles()
        pb.visible = True
        progress_state.value = "正在写入文件，稍后..."
        router.page.update()
        sleep(1)
        progress_state.value = "写入文件完成，请检查"
        pb.value = 1
        router.page.update()

    get_directory_dialog = ft.FilePicker(on_result=get_directory_result)

    router.page.overlay.extend([get_directory_dialog])
    router.page.add(
        ft.Container(
            margin=ft.margin.only(top=20),
            content=ft.Row(
                [
                    ft.FilledButton("返回首页", icon="ARROW_BACK_IOS_OUTLINED", on_click=goback),
                    ft.Card(
                        width=600,
                        height=450,
                        content=
                        ft.Container(
                            image_src=f"../assets/bg_content.jpeg",
                            image_fit=ft.ImageFit.COVER,
                            border_radius=5,
                            content=ft.Column(
                                [
                                    ft.Container(
                                        width=600,
                                        height=50,
                                        alignment=ft.alignment.center,
                                        content=ft.FilledButton(
                                            "打开文件夹",
                                            icon=ft.icons.FOLDER_OPEN,
                                            on_click=lambda _: get_directory_dialog.get_directory_path(),
                                        ),
                                    ),
                                    ft.Container(
                                        width=600,
                                        height=70,
                                        padding=10,
                                        alignment=ft.alignment.center,
                                        content=directory_path,
                                    ),
                                    ft.FilledButton("开始写入", icon="DOWNLOADING", on_click=write_file),
                                    ft.Container(
                                        width=600,
                                        height=50,
                                        alignment=ft.alignment.center,
                                        content=ft.Column([progress_state, pb]),
                                    ),

                                ],
                                wrap=True,
                                spacing=30,
                                run_spacing=10,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            )
                        )),
                ],
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.START
            )
        ),
    )
